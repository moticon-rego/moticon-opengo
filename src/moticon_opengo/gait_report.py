import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Union

import numpy as np
from openpyxl import load_workbook

from moticon_opengo.utils import FileIterator, Side, Step, is_not_null


class GaitReportFileIterator(FileIterator):
    def __init__(self, folder: str):
        super().__init__(folder, ".xlsx")

    def _check_files(self):
        """
        Plausibility check removing files which are obviously no OpenGo gait report
        files.
        """
        for fname in self.files:
            wb = load_workbook(fname)
            required_sheet_names = ["Report Data", "Left Steps", "Right Steps"]

            if not set(required_sheet_names).issubset(set(wb.sheetnames)):
                self.files.remove(fname)

            wb.close()


class GaitReportStep(Step):
    def __init__(self, side: Side):
        super().__init__(side)
        self._field_names: Dict[str, str] = list()
        self.abs_time: Optional[datetime] = None

        self.sequence_number: Optional[int] = None
        self.time_initial_contact: Optional[float] = None
        self.time_toe_off: Optional[float] = None
        self.gait_line_start_x: Optional[float] = None
        self.gait_line_start_y: Optional[float] = None
        self.gait_line_end_x: Optional[float] = None
        self.gait_line_end_y: Optional[float] = None
        self.gait_line_length: Optional[float] = None
        self.gait_line_width: Optional[float] = None
        self.first_peak: Optional[float] = None
        self.first_peak_timing: Optional[float] = None
        self.second_peak: Optional[float] = None
        self.second_peak_timing: Optional[float] = None
        self.local_minimum: Optional[float] = None
        self.local_minimum_timing: Optional[float] = None
        self.max_tf: Optional[float] = None
        self.mean_tf: Optional[float] = None
        self.stance_time: Optional[float] = None
        self.swing_time: Optional[float] = None
        self.double_support: Optional[float] = None
        self.double_support_timing: Optional[float] = None
        self.step_duration: Optional[float] = None
        self.stride_length: Optional[float] = None
        self.stride_time: Optional[float] = None
        self.stride_velocity: Optional[float] = None
        self.stride_cadence: Optional[float] = None

        self._define_field_names()

    def _define_field_names(self):
        self._field_names = {
            "Step": "sequence_number",
            "Initial contact (s)": "time_initial_contact",
            "Toe off (s)": "time_toe_off",
            "Gaitline start x": "gait_line_start_x",
            "Gaitline start x (%)": "gait_line_start_x",
            "Gaitline start y": "gait_line_start_y",
            "Gaitline start y (%)": "gait_line_start_y",
            "Gaitline end x": "gait_line_end_x",
            "Gaitline end x (%)": "gait_line_end_x",
            "Gaitline end y": "gait_line_end_y",
            "Gaitline end y (%)": "gait_line_end_y",
            "Gaitline length": "gait_line_length",
            "Gaitline length (%)": "gait_line_length",
            "Gaitline width": "gait_line_width",
            "Gaitline width (%)": "gait_line_width",
            "First peak (N)": "first_peak",
            "First peak timing (as % of stance)": "first_peak_timing",
            "Second peak (N)": "second_peak",
            "Second peak timing (as % of stance)": "second_peak_timing",
            "Local minimum between peaks (N)": "local_minimum",
            "Local minimum timing (as % of stance)": "local_minimum_timing",
            "Max total force during stance phase (N)": "max_tf",
            "Mean total force during stance phase (N)": "mean_tf",
            "Stance (s)": "stance_time",
            "Swing (s)": "swing_time",
            "Double support (s)": "double_support",
            "Double support (as % of stance)": "double_support_timing",
            "Step duration (s)": "step_duration",
            "Stride length (m)": "stride_length",
            "Stride time (s)": "stride_time",
            "Stride velocity (m/s)": "stride_velocity",
            "Stride cadence (1/min)": "stride_cadence",
        }

    def set_data(self, values: List[Union[int, float]], headers: List[str]):
        deprecated_field_names: List[str] = [
            "Gaitline start x (mm)",
            "Gaitline start y (mm)",
            "Gaitline end x (mm)",
            "Gaitline end y (mm)",
            "Gaitline length (mm)",
            "Gaitline width (mm)",
        ]

        for v, h in zip(values, headers):
            try:
                setattr(self, self._field_names[h], v)
            except KeyError:
                if h not in deprecated_field_names:
                    raise KeyError(f'The header entry "{h}" is unknown')

    @property
    def max_peak(self) -> Optional[float]:
        """
        From the first and second total force peak, return the larger one.
        """
        peaks: List[float] = [
            x for x in [self.first_peak, self.second_peak] if is_not_null(x)
        ]

        if peaks:
            return max(peaks)

        return None


class GaitReport(object):
    def __init__(
        self,
        fname: str,
        start_idx: int = -20,
        stop_idx: int = -5,
        datetime_format: str = "%y%m%d_%H_%M_%S",
    ):
        """
        Represents a gait report result constructed from a spreadsheet export
        file [fname]. If the file basename character range [start_idx:stop_idx]
        contains the absolute start time of the data, then it can be parsed
        using [datetime_format]. The steps will then not only have a time
        relative to the measurement start, but also an absolute datetime.
        """
        self.fname: str = fname
        self.abs_start_time: Optional[datetime] = None

        try:
            self.abs_start_time = datetime.strptime(
                os.path.basename(fname)[start_idx:stop_idx], datetime_format
            )
        except ValueError:
            pass

        self._field_names: Dict[str, str] = list()
        self._step_count: Optional[int] = None
        self._step_count_statistics: Optional[int] = None
        self._gait_line_dims_left: Optional[np.array] = None
        self._gait_line_dims_right: Optional[np.array] = None
        self._gait_line_start_point_left: Optional[np.array] = None
        self._gait_line_start_point_right: Optional[np.array] = None
        self._gait_line_end_point_left: Optional[np.array] = None
        self._gait_line_end_point_right: Optional[np.array] = None
        self._gait_line_start_point_std_left: Optional[np.array] = None
        self._gait_line_start_point_std_right: Optional[np.array] = None
        self._gait_line_end_point_std_left: Optional[np.array] = None
        self._gait_line_end_point_std_right: Optional[np.array] = None
        self._gait_line_ap_left: Optional[np.array] = None
        self._gait_line_ml_left: Optional[np.array] = None
        self._gait_line_ml_std_left: Optional[np.array] = None
        self._gait_line_ap_right: Optional[np.array] = None
        self._gait_line_ml_right: Optional[np.array] = None
        self._gait_line_ml_std_right: Optional[np.array] = None
        self._mean_pressure_left: Optional[np.array] = None
        self._mean_pressure_right: Optional[np.array] = None
        self._max_pressure_left: Optional[np.array] = None
        self._max_pressure_right: Optional[np.array] = None
        self._mean_pressure_stance_left: Optional[np.array] = None
        self._mean_pressure_stance_right: Optional[np.array] = None
        self._max_pressure_stance_left: Optional[np.array] = None
        self._max_pressure_stance_right: Optional[np.array] = None
        self._mean_pressure_initial_contact_left: Optional[np.array] = None
        self._mean_pressure_initial_contact_right: Optional[np.array] = None
        self._mean_pressure_mid_stance_left: Optional[np.array] = None
        self._mean_pressure_mid_stance_right: Optional[np.array] = None
        self._mean_pressure_terminal_stance_left: Optional[np.array] = None
        self._mean_pressure_terminal_stance_right: Optional[np.array] = None
        self._mean_tf_stance_left: Optional[np.array] = None
        self._mean_tf_stance_right: Optional[np.array] = None
        self._max_tf_stance_left: Optional[np.array] = None
        self._max_tf_stance_right: Optional[np.array] = None
        self._mean_max_tf_stance_left: Optional[np.array] = None
        self._mean_max_tf_stance_right: Optional[np.array] = None
        self._total_force_left: Optional[np.array] = None
        self._total_force_right: Optional[np.array] = None
        self._total_force_std_left: Optional[np.array] = None
        self._total_force_std_right: Optional[np.array] = None
        self._cycle_time: Optional[np.array] = None
        self._cadence: Optional[np.array] = None
        self._double_support_time: Optional[np.array] = None
        self._double_support_fraction: Optional[np.array] = None
        self._double_support_time_left: Optional[np.array] = None
        self._double_support_time_right: Optional[np.array] = None
        self._step_duration_left: Optional[np.array] = None
        self._step_duration_right: Optional[np.array] = None
        self._stance_duration_left: Optional[np.array] = None
        self._stance_duration_right: Optional[np.array] = None
        self._stance_duration_std_left: Optional[np.array] = None
        self._stance_duration_std_right: Optional[np.array] = None
        self._swing_duration_left: Optional[np.array] = None
        self._swing_duration_right: Optional[np.array] = None
        self._swing_duration_std_left: Optional[np.array] = None
        self._swing_duration_std_right: Optional[np.array] = None
        self._stance_fraction_left: Optional[np.array] = None
        self._stance_fraction_right: Optional[np.array] = None
        self._stance_fraction_std_left: Optional[np.array] = None
        self._stance_fraction_std_right: Optional[np.array] = None
        self._swing_fraction_left: Optional[np.array] = None
        self._swing_fraction_right: Optional[np.array] = None
        self._swing_fraction_std_left: Optional[np.array] = None
        self._swing_fraction_std_right: Optional[np.array] = None
        self._weight_histogram_left: Optional[np.array] = None
        self._weight_histogram_right: Optional[np.array] = None
        self._weight_histogram_bins: Optional[np.array] = None
        self._weight_histogram_stance_left: Optional[np.array] = None
        self._weight_histogram_stance_right: Optional[np.array] = None
        self._weight_histogram_stance_bins: Optional[np.array] = None
        self._acceleration_x_left: Optional[np.array] = None
        self._acceleration_y_left: Optional[np.array] = None
        self._acceleration_z_left: Optional[np.array] = None
        self._acceleration_std_x_left: Optional[np.array] = None
        self._acceleration_std_y_left: Optional[np.array] = None
        self._acceleration_std_z_left: Optional[np.array] = None
        self._acceleration_x_right: Optional[np.array] = None
        self._acceleration_y_right: Optional[np.array] = None
        self._acceleration_z_right: Optional[np.array] = None
        self._acceleration_std_x_right: Optional[np.array] = None
        self._acceleration_std_y_right: Optional[np.array] = None
        self._acceleration_std_z_right: Optional[np.array] = None
        self._stride_length: Optional[np.array] = None
        self._distance: Optional[np.array] = None
        self._speed: Optional[np.array] = None

        self.steps: List[GaitReportStep] = list()

        self._define_field_names()
        self._load_data()

    def __str__(self) -> str:
        return f"OpenGo Gait Report <{self.fname}>"

    def _load_data(self):
        self._load_report()
        self._load_steps()

    def _load_report(self):
        wb = load_workbook(self.fname)
        ws = wb.get_sheet_by_name("Report Data")

        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] in self._field_names:
                data = [x for x in row[1:] if x is not None]

                if not data:
                    data = None
                elif len(data) == 1:
                    data = data[0]
                else:
                    data = np.array(data)

                setattr(self, self._field_names[row[0]], data)

    def _load_steps(self):
        steps: List[GaitReportStep] = list()

        for side in Side:
            wb = load_workbook(self.fname)
            ws = wb.get_sheet_by_name(f"{side.name.capitalize()} Steps")
            headers: List[str] = list()

            for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
                if i == 0:
                    headers = row
                    continue

                steps.append(GaitReportStep(side))
                steps[-1].set_data(row, headers)

                if self.abs_start_time:
                    steps[-1].abs_time = self.abs_start_time + timedelta(
                        seconds=steps[-1].time_initial_contact
                    )

        self.steps = sorted(steps, key=lambda x: x.time_initial_contact)

    @property
    def has_absolute_time(self) -> bool:
        return self.abs_start_time is not None

    @property
    def calendar_days_with_steps(self):
        if not self.has_absolute_time:
            return list()

        unique_calendar_days = set()

        for dt in [step.abs_time for step in self.steps]:
            calendar_day = dt.date()
            unique_calendar_days.add(calendar_day)

        return list(sorted(list(unique_calendar_days)))

    def steps_of_calendar_day(self, calendar_day):
        if not self.has_absolute_time:
            return list()

        steps = list(self.steps)
        return [step for step in steps if step.abs_time.date() == calendar_day]

    @property
    def field_names(self):
        return self._field_names

    @property
    def step_count(self):
        return self._step_count

    @property
    def step_count_statistics(self):
        return self._step_count_statistics

    @property
    def gait_line_dims(self):
        return {
            Side.LEFT: self._gait_line_dims_left,
            Side.RIGHT: self._gait_line_dims_right,
        }

    @property
    def gait_line_start_point(self):
        return {
            Side.LEFT: self._gait_line_start_point_left,
            Side.RIGHT: self._gait_line_start_point_right,
        }

    @property
    def gait_line_end_point(self):
        return {
            Side.LEFT: self._gait_line_end_point_left,
            Side.RIGHT: self._gait_line_end_point_right,
        }

    @property
    def gait_line_start_point_std(self):
        return {
            Side.LEFT: self._gait_line_start_point_std_left,
            Side.RIGHT: self._gait_line_start_point_std_right,
        }

    @property
    def gait_line_end_point_std(self):
        return {
            Side.LEFT: self._gait_line_end_point_std_left,
            Side.RIGHT: self._gait_line_end_point_std_right,
        }

    @property
    def gait_line(self):
        return {
            Side.LEFT: np.column_stack(
                (self._gait_line_ap_left, self._gait_line_ml_left)
            ),
            Side.RIGHT: np.column_stack(
                (self._gait_line_ap_left, self._gait_line_ml_left)
            ),
        }

    @property
    def gait_line_std(self):
        return {
            Side.LEFT: self._gait_line_ml_std_left,
            Side.RIGHT: self._gait_line_ml_std_right,
        }

    @property
    def mean_pressure(self):
        return {
            Side.LEFT: self._mean_pressure_left,
            Side.RIGHT: self._mean_pressure_right,
        }

    @property
    def max_pressure(self):
        return {
            Side.LEFT: self._max_pressure_left,
            Side.RIGHT: self._max_pressure_right,
        }

    @property
    def mean_pressure_stance(self):
        return {
            Side.LEFT: self._mean_pressure_stance_left,
            Side.RIGHT: self._mean_pressure_stance_right,
        }

    @property
    def max_pressure_stance(self):
        return {
            Side.LEFT: self._max_pressure_stance_left,
            Side.RIGHT: self._max_pressure_stance_right,
        }

    @property
    def mean_pressure_initial_contact(self):
        return {
            Side.LEFT: self._mean_pressure_initial_contact_left,
            Side.RIGHT: self._mean_pressure_initial_contact_right,
        }

    @property
    def mean_pressure_mid_stance(self):
        return {
            Side.LEFT: self._mean_pressure_mid_stance_left,
            Side.RIGHT: self._mean_pressure_mid_stance_right,
        }

    @property
    def mean_pressure_terminal_stance(self):
        return {
            Side.LEFT: self._mean_pressure_terminal_stance_left,
            Side.RIGHT: self._mean_pressure_terminal_stance_right,
        }

    @property
    def mean_tf_stance(self):
        return {
            Side.LEFT: self._mean_tf_stance_left,
            Side.RIGHT: self._mean_tf_stance_right,
        }

    @property
    def max_tf_stance(self):
        return {
            Side.LEFT: self._max_tf_stance_left,
            Side.RIGHT: self._max_tf_stance_right,
        }

    @property
    def mean_max_tf_stance(self):
        return {
            Side.LEFT: self._mean_max_tf_stance_left,
            Side.RIGHT: self._mean_max_tf_stance_right,
        }

    @property
    def total_force(self):
        return {
            Side.LEFT: self._total_force_left,
            Side.RIGHT: self._total_force_right,
        }

    @property
    def total_force_std(self):
        return {
            Side.LEFT: self._total_force_std_left,
            Side.RIGHT: self._total_force_std_right,
        }

    @property
    def cycle_time(self):
        return self._cycle_time

    @property
    def cadence(self):
        return self._cadence

    @property
    def double_support_time(self):
        return self._double_support_time

    @property
    def double_support_fraction(self):
        return self._double_support_fraction

    @property
    def double_support_time_side(self):
        return {
            Side.LEFT: self._double_support_time_left,
            Side.RIGHT: self._double_support_time_right,
        }

    @property
    def step_duration(self):
        return {
            Side.LEFT: self._step_duration_left,
            Side.RIGHT: self._step_duration_right,
        }

    @property
    def stance_duration(self):
        return {
            Side.LEFT: self._stance_duration_left,
            Side.RIGHT: self._stance_duration_right,
        }

    @property
    def stance_duration_std(self):
        return {
            Side.LEFT: self._stance_duration_std_left,
            Side.RIGHT: self._stance_duration_std_right,
        }

    @property
    def swing_duration(self):
        return {
            Side.LEFT: self._swing_duration_left,
            Side.RIGHT: self._swing_duration_right,
        }

    @property
    def swing_duration_std(self):
        return {
            Side.LEFT: self._swing_duration_std_left,
            Side.RIGHT: self._swing_duration_std_right,
        }

    @property
    def stance_fraction(self):
        return {
            Side.LEFT: self._stance_fraction_left,
            Side.RIGHT: self._stance_fraction_right,
        }

    @property
    def stance_fraction_std(self):
        return {
            Side.LEFT: self._stance_fraction_std_left,
            Side.RIGHT: self._stance_fraction_std_right,
        }

    @property
    def swing_fraction(self):
        return {
            Side.LEFT: self._swing_fraction_left,
            Side.RIGHT: self._swing_fraction_right,
        }

    @property
    def swing_fraction_std(self):
        return {
            Side.LEFT: self._swing_fraction_std_left,
            Side.RIGHT: self._swing_fraction_std_right,
        }

    @property
    def weight_histogram(self):
        return {
            Side.LEFT: self._weight_histogram_left,
            Side.RIGHT: self._weight_histogram_right,
        }

    @property
    def weight_histogram_bins(self):
        return self._weight_histogram_bins

    @property
    def weight_histogram_stance(self):
        return {
            Side.LEFT: self._weight_histogram_stance_left,
            Side.RIGHT: self._weight_histogram_stance_right,
        }

    @property
    def weight_histogram_stance_bins(self):
        return self._weight_histogram_stance_bins

    @property
    def acceleration(self):
        return {
            Side.LEFT: np.column_stack(
                (
                    self._acceleration_x_left,
                    self._acceleration_y_left,
                    self._acceleration_z_left,
                )
            ),
            Side.RIGHT: np.column_stack(
                (
                    self._acceleration_x_right,
                    self._acceleration_y_right,
                    self._acceleration_z_right,
                )
            ),
        }

    @property
    def acceleration_std(self):
        return {
            Side.LEFT: np.column_stack(
                (
                    self._acceleration_std_x_left,
                    self._acceleration_std_y_left,
                    self._acceleration_std_z_left,
                )
            ),
            Side.RIGHT: np.column_stack(
                (
                    self._acceleration_std_x_right,
                    self._acceleration_std_y_right,
                    self._acceleration_std_z_right,
                )
            ),
        }

    @property
    def stride_length(self):
        return self._stride_length

    @property
    def distance(self):
        return self._distance

    @property
    def speed(self):
        return self._speed

    def _define_field_names(self):
        self._field_names = {
            "Number of steps": "_step_count",
            "Number of steps used for statistics": "_step_count_statistics",
            "Mean length/width of gait line (left)": "_gait_line_dims_left",
            "Mean length/width of gait line (right)": "_gait_line_dims_right",
            "Mean startpoint x/y of gait line (left)": "_gait_line_start_point_left",
            "Mean startpoint x/y of gait line (right)": "_gait_line_start_point_right",
            "Mean endpoint x/y of gait line (left)": "_gait_line_end_point_left",
            "Mean endpoint x/y of gait line (right)": "_gait_line_end_point_right",
            "Standard deviation x/y of gait line startpoint (left)": "_gait_line_start_point_std_left",
            "Standard deviation x/y of gait line startpoint (right)": "_gait_line_start_point_std_right",
            "Standard deviation x/y of gait line endpoint (left)": "_gait_line_end_point_std_left",
            "Standard deviation x/y of gait line endpoint (right)": "_gait_line_end_point_std_right",
            "Mean gait line (anterior-posterior) (left)": "_gait_line_ap_left",
            "Mean gait line (medial-lateral) (left)": "_gait_line_ml_left",
            "Stddev of average gait line (medial-lateral) (left)": "_gait_line_ml_std_left",
            "Mean gait line (anterior-posterior) (right)": "_gait_line_ap_right",
            "Mean gait line (medial-lateral) (right)": "_gait_line_ml_right",
            "Stddev of average gait line (medial-lateral) (right)": "_gait_line_ml_std_right",
            "Mean (left) (N/cm²)": "_mean_pressure_left",
            "Mean (right) (N/cm²)": "_mean_pressure_right",
            "Maximum (left) (N/cm²)": "_max_pressure_left",
            "Maximum (right) (N/cm²)": "_max_pressure_right",
            "Mean during stance phase (left) (N/cm²)": "_mean_pressure_stance_left",
            "Mean during stance phase (right) (N/cm²)": "_mean_pressure_stance_right",
            "Maximum during stance phase (left) (N/cm²)": "_max_pressure_stance_left",
            "Maximum during stance phase (right) (N/cm²)": "_max_pressure_stance_right",
            "During initial contact (left) (N/cm²)": "_mean_pressure_initial_contact_left",
            "During initial contact (right) (N/cm²)": "_mean_pressure_initial_contact_right",
            "During mid stance (left) (N/cm²)": "_mean_pressure_mid_stance_left",
            "During mid stance (right) (N/cm²)": "_mean_pressure_mid_stance_right",
            "During terminal stance (left) (N/cm²)": "_mean_pressure_terminal_stance_left",
            "During terminal stance (right) (N/cm²)": "_mean_pressure_terminal_stance_right",
            "Mean total force during stance phase (left) (N)": "_mean_tf_stance_left",
            "Mean total force during stance phase (right) (N)": "_mean_tf_stance_right",
            "Maximum total force during stance phase (left) (N)": "_max_tf_stance_left",
            "Maximum total force during stance phase (right) (N)": "_max_tf_stance_right",
            "Mean of all Maxima of total force during all stance phases (left) (N)": "_mean_max_tf_stance_left",
            "Mean of all Maxima of total force during all stance phases (right) (N)": "_mean_max_tf_stance_right",
            "Mean total force curve (left) (N)": "_total_force_left",
            "Mean total force curve (right) (N)": "_total_force_right",
            "Stddev of total force curve (left) (N)": "_total_force_std_left",
            "Stddev of total force curve (right) (N)": "_total_force_std_right",
            "Mean gait cycle time (s)": "_gait_cycle_time",
            "Mean gait cadence (strides per minute)": "_cadence",
            "Mean double support time (s)": "_double_support_time",
            "Mean fraction of double support (%)": "_double_support_fraction",
            "Mean double support time (left) (s)": "_double_support_time_left",
            "Mean double support time (right) (s)": "_double_support_time_right",
            "Mean step duration (left) (s)": "_step_duration_left",
            "Mean step duration (right) (s)": "_step_duration_right",
            "Mean stance duration (left) (s)": "_stance_duration_left",
            "Mean stance duration (right) (s)": "_stance_duration_right",
            "Stddev of stance duration (left) (s)": "_stance_duration_std_left",
            "Stddev of stance duration (right) (s)": "_stance_duration_std_right",
            "Mean swing duration (left) (s)": "_swing_duration_left",
            "Mean swing duration (right) (s)": "_swing_duration_right",
            "Stddev of swing duration (left) (s)": "_swing_duration_std_left",
            "Stddev of swing duration (right) (s)": "_swing_duration_std_right",
            "Mean fraction of stance phase (left) (%)": "_stance_fraction_left",
            "Mean fraction of stance phase (right) (%)": "_stance_fraction_right",
            "Stddev of fraction of stance phase (left) (%)": "_stance_fraction_std_left",
            "Stddev of fraction of stance phase (right) (%)": "_stance_fraction_std_right",
            "Mean fraction of swing phase (left) (%)": "_swing_fraction_left",
            "Mean fraction of swing phase (right) (%)": "_swing_fraction_right",
            "Stddev of fraction of swing phase (left) (%)": "_swing_fraction_std_left",
            "Stddev of fraction of swing phase (right) (%)": "_swing_fraction_std_right",
            "Weight histogram values (left) (%)": "_weight_histogram_left",
            "Weight histogram values (right) (%)": "_weight_histogram_right",
            "Weight histogram bins (kg)": "_weight_histogram_bins",
            "Weight histogram values during stance phase (left) (%)": "_weight_histogram_stance_left",
            "Weight histogram values during stance phase (right) (%)": "_weight_histogram_stance_right",
            "Weight histogram bins during stance (kg)": "_weight_histogram_stance_bins",
            "Mean acceleration (x) over gait cycle (left) (g)": "_acceleration_x_left",
            "Mean acceleration (y) over gait cycle (left) (g)": "_acceleration_y_left",
            "Mean acceleration (z) over gait cycle (left) (g)": "_acceleration_z_left",
            "Stddev of acceleration (x) over gait cycle (left) (g)": "_acceleration_std_x_left",
            "Stddev of acceleration (y) over gait cycle (left) (g)": "_acceleration_std_y_left",
            "Stddev of acceleration (z) over gait cycle (left) (g)": "_acceleration_std_z_left",
            "Mean acceleration (x) over gait cycle (right) (g)": "_acceleration_x_right",
            "Mean acceleration (y) over gait cycle (right) (g)": "_acceleration_y_right",
            "Mean acceleration (z) over gait cycle (right) (g)": "_acceleration_z_right",
            "Stddev of acceleration (x) over gait cycle (right) (g)": "_acceleration_std_x_right",
            "Stddev of acceleration (y) over gait cycle (right) (g)": "_acceleration_std_y_right",
            "Stddev of acceleration (z) over gait cycle (right) (g)": "_acceleration_std_z_right",
            "Mean stride length (m)": "_stride_length",
            "Walking distance (m)": "_distance",
            "Mean walking speed (m/s)": "_speed",
        }
